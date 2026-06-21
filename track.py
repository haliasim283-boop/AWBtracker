import json
import time
import tkinter as tk
from tkinter import filedialog

import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
USERNAME = "hamza"
PASSWORD = "hamza1234"
# ─────────────────────────────────────────────

LOGIN_URL    = "https://cargospot-portal.champ.aero/index.asp?portal_id=PIA"
TRACKING_URL = "https://cargospot-portal.champ.aero/tracking.asp"

# Columns to write portal results into (1-indexed: Q=17 … U=21)
OUT_COLS = {
    "uorigin":      17,   # Q
    "udest":        18,   # R
    "uweight":      19,   # S
    "upieces":      20,   # T
    "ustatus":      21,   # U
}

PINK_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")


def pick_excel_file():
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Select the AWB Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
    )
    root.destroy()
    return path


def load_awbs_from_excel(file_path, start_row, end_row):
    """
    Reads rows [start_row..end_row], returns a list of dicts for every row
    whose PREFIX starts with '214-'.

    PREFIX "214-8220" + AWB "2223"  ->  serial "82202223"
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    # Locate PREFIX and AWB columns from the header row (row 1)
    prefix_col = awb_col = None
    for cell in ws[1]:
        if cell.value is None:
            continue
        h = str(cell.value).strip().upper()
        if h == "PREFIX":
            prefix_col = cell.column
        elif h == "AWB":
            awb_col = cell.column

    if prefix_col is None or awb_col is None:
        wb.close()
        raise ValueError(
            f"Could not find 'PREFIX' or 'AWB' header in row 1 "
            f"(prefix_col={prefix_col}, awb_col={awb_col})."
        )

    awbs = []
    for row_num in range(start_row, end_row + 1):
        prefix_val = ws.cell(row=row_num, column=prefix_col).value
        awb_val    = ws.cell(row=row_num, column=awb_col).value

        if not prefix_val:
            continue
        prefix_str = str(prefix_val).strip()
        if not prefix_str.upper().startswith("214-"):
            continue

        if awb_val is None:
            print(f"  Row {row_num}: PREFIX={prefix_str} — AWB cell empty, skipped.")
            continue

        try:
            awb_str = str(int(float(str(awb_val).strip())))
        except ValueError:
            print(f"  Row {row_num}: cannot parse AWB '{awb_val}', skipped.")
            continue

        suffix = prefix_str[4:]          # "214-8220" -> "8220"
        serial = suffix + awb_str        # "8220" + "2223" -> "82202223"
        awbs.append({"prefix": "214", "serial": serial, "row": row_num})
        print(f"  Row {row_num}: {prefix_str} + {awb_str}  ->  serial={serial}")

    wb.close()
    return awbs


def write_results_to_excel(file_path, results):
    """Write portal results back to columns Q-U in the source Excel file."""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Write / ensure column headers in row 1
    for field, col in OUT_COLS.items():
        cell = ws.cell(row=1, column=col)
        if not cell.value:
            cell.value = field

    written = 0
    for entry in results:
        if entry.get("status") != "success":
            continue

        row_num = entry["excel_row"]
        data    = entry["data"]

        field_map = {
            "uorigin":  data.get("origin", ""),
            "udest":    data.get("destination", ""),
            "uweight":  data.get("weight", ""),
            "upieces":  data.get("pieces", ""),
            "ustatus":  data.get("status", ""),
        }

        for field, col in OUT_COLS.items():
            cell       = ws.cell(row=row_num, column=col)
            cell.value = field_map[field]
            cell.fill  = PINK_FILL

        written += 1

    wb.save(file_path)
    print(f"\nWrote portal results for {written} AWB(s) back to: {file_path}")


def login(browser):
    print("Navigating to login page...")
    browser.get(LOGIN_URL)
    wait = WebDriverWait(browser, 60)
    username_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#user_name")))
    browser.execute_script("arguments[0].value = arguments[1];", username_field, USERNAME)
    password_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#password")))
    browser.execute_script("arguments[0].value = arguments[1];", password_field, PASSWORD)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#btn-login"))).click()
    wait.until(EC.url_changes(LOGIN_URL))
    print("Login successful.\n")


def fetch_awb(browser, prefix, serial):
    wait = WebDriverWait(browser, 60)

    browser.get(TRACKING_URL)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#awb_1")))
    time.sleep(1)

    awb_input = browser.find_element(By.CSS_SELECTOR, "#awb_1")
    browser.execute_script("arguments[0].value = arguments[1];", awb_input, serial)
    browser.execute_script(
        "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", awb_input
    )
    print(f"  Entered serial: {serial}")

    submit_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#track-submit")))
    browser.execute_script("arguments[0].click();", submit_btn)
    print("  Submit clicked, waiting for AJAX to complete...")

    try:
        WebDriverWait(browser, 5).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "#overlay"))
        )
    except Exception:
        pass

    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#overlay")))
    print("  Overlay gone — results ready.")
    time.sleep(0.5)

    return scrape_results(browser, prefix, serial)


def get_span(browser, css_class):
    try:
        els = browser.find_elements(By.CSS_SELECTOR, f"span.{css_class}")
        if els:
            text = els[0].text.strip()
            return text.split(":", 1)[1].strip() if ":" in text else text
    except Exception:
        pass
    return ""


def scrape_results(browser, prefix, serial):
    full_awb = f"{prefix}-{serial}"
    origin      = get_span(browser, "tracking_origin")
    destination = get_span(browser, "tracking_destination")
    pieces      = get_span(browser, "tracking_pieces")
    weight      = get_span(browser, "tracking_weight")
    status      = get_span(browser, "tracking_status")
    flight_no   = (get_span(browser, "tracking_flight")
                   or get_span(browser, "tracking_flight_no")
                   or get_span(browser, "tracking_flightno"))
    date        = (get_span(browser, "tracking_date")
                   or get_span(browser, "tracking_flight_date"))

    return {
        "awb":         full_awb,
        "origin":      origin,
        "destination": destination,
        "pieces":      pieces,
        "weight":      weight,
        "status":      status,
        "flight_no":   flight_no,
        "date":        date,
    }


def main():
    print("\n── PIA CargoSpot Selenium Tracker ──\n")

    # ── Step 1: pick Excel file ──────────────────────────────────────────────
    print("Please select the Excel file in the dialog that opens...")
    excel_path = pick_excel_file()
    if not excel_path:
        print("No file selected. Exiting.")
        return
    print(f"Selected: {excel_path}\n")

    # ── Step 2: row range ────────────────────────────────────────────────────
    while True:
        try:
            start_row = int(input("Enter START row number (e.g. 700): ").strip())
            end_row   = int(input("Enter END   row number (e.g. 740): ").strip())
            if start_row < 2:
                print("Start row must be 2 or higher (row 1 is the header).")
                continue
            if end_row < start_row:
                print("End row must be >= start row.")
                continue
            break
        except ValueError:
            print("Please enter valid integers.")

    # ── Step 3: extract PIA AWBs ─────────────────────────────────────────────
    print(f"\nScanning rows {start_row}–{end_row} for PIA (214-xxxx) AWBs...\n")
    try:
        target_awbs = load_awbs_from_excel(excel_path, start_row, end_row)
    except Exception as e:
        print(f"Failed to read Excel: {e}")
        return

    if not target_awbs:
        print("No PIA AWBs found in the specified range. Exiting.")
        return

    print(f"\nFound {len(target_awbs)} PIA AWB(s) to track.\n")

    # ── Step 4: launch browser and track ─────────────────────────────────────
    options = webdriver.FirefoxOptions()
    browser = webdriver.Firefox(options=options)
    browser.maximize_window()

    results = []

    try:
        login(browser)

        for i, entry in enumerate(target_awbs, 1):
            label = f"214-{entry['serial']}  (row {entry['row']})"
            print(f"[{i}/{len(target_awbs)}] Fetching {label}...")
            try:
                data = fetch_awb(browser, entry["prefix"], entry["serial"])
                results.append({"status": "success", "excel_row": entry["row"], "data": data})
                print(f"  origin={data['origin']}  dest={data['destination']}  "
                      f"pcs={data['pieces']}  wt={data['weight']}  status={data['status']}\n")
            except Exception as e:
                print(f"  FAILED: {e}\n")
                results.append({
                    "awb":       f"214-{entry['serial']}",
                    "excel_row": entry["row"],
                    "status":    "error",
                    "error":     str(e),
                })

            time.sleep(1)

    except Exception as e:
        print(f"Fatal error: {e}")

    finally:
        # ── Step 5: save JSON ────────────────────────────────────────────────
        print("\n── Results ──\n")
        print(json.dumps(results, indent=2))
        with open("awb_results.json", "w") as f:
            json.dump(results, f, indent=2)
        print("Saved to: awb_results.json")

        # ── Step 6: write results back to Excel ─────────────────────────────
        try:
            write_results_to_excel(excel_path, results)
        except Exception as e:
            print(f"Could not write results to Excel: {e}")

        input("\nPress Enter to close the browser...")
        browser.quit()


if __name__ == "__main__":
    main()
