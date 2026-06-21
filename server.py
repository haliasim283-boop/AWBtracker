import json
import os
import shutil
import time
import threading
import uuid

import openpyxl
from openpyxl.styles import PatternFill
from flask import Flask, request, jsonify, Response, send_file, stream_with_context
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ─────────────────────────────────────────────
USERNAME = "hamza"
PASSWORD = "hamza1234"
# ─────────────────────────────────────────────

LOGIN_URL    = "https://cargospot-portal.champ.aero/index.asp?portal_id=PIA"
TRACKING_URL = "https://cargospot-portal.champ.aero/tracking.asp"

OUT_COLS = {
    "uorigin":  17,   # Q
    "udest":    18,   # R
    "uweight":  19,   # S
    "upieces":  20,   # T
    "ustatus":  21,   # U
}

PINK_FILL = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

app = Flask(__name__)

# job_id -> {"status": "running"|"done"|"error", "log": [...], "result_path": str|None, "error": str|None}
jobs = {}


# ── Excel helpers ─────────────────────────────────────────────────────────────

def load_awbs_from_excel(file_path, start_row, end_row, log):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    prefix_col = awb_col = None
    for cell in ws[1]:
        if cell.value is None:
            continue
        h = str(cell.value).strip().upper()
        if h == "PREFIX":
            prefix_col = cell.column
        elif h == "AWB":
            awb_col = cell.column

    if prefix_col is None or awb_col is None:
        wb.close()
        raise ValueError("Could not find 'PREFIX' or 'AWB' header in row 1.")

    awbs = []
    for row_num in range(start_row, end_row + 1):
        prefix_val = ws.cell(row=row_num, column=prefix_col).value
        awb_val    = ws.cell(row=row_num, column=awb_col).value

        if not prefix_val:
            continue
        prefix_str = str(prefix_val).strip()
        if not prefix_str.upper().startswith("214-"):
            continue

        if awb_val is None:
            log(f"Row {row_num}: PREFIX={prefix_str} — AWB cell empty, skipped.")
            continue

        try:
            awb_str = str(int(float(str(awb_val).strip())))
        except ValueError:
            log(f"Row {row_num}: Cannot parse AWB '{awb_val}', skipped.")
            continue

        suffix = prefix_str[4:]          # "214-8220" -> "8220"
        serial = suffix + awb_str        # "8220" + "2223" -> "82202223"
        log(f"Row {row_num}: {prefix_str} + {awb_str}  ->  serial={serial}")
        awbs.append({"prefix": "214", "serial": serial, "row": row_num})

    wb.close()
    return awbs


def write_results_to_excel(file_path, results, log):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for field, col in OUT_COLS.items():
        cell = ws.cell(row=1, column=col)
        if not cell.value:
            cell.value = field

    written = 0
    for entry in results:
        if entry.get("status") != "success":
            continue
        row_num = entry["excel_row"]
        data    = entry["data"]
        field_map = {
            "uorigin":  data.get("origin", ""),
            "udest":    data.get("destination", ""),
            "uweight":  data.get("weight", ""),
            "upieces":  data.get("pieces", ""),
            "ustatus":  data.get("status", ""),
        }
        for field, col in OUT_COLS.items():
            cell = ws.cell(row=row_num, column=col)
            cell.value = field_map[field]
            cell.fill  = PINK_FILL
        written += 1

    wb.save(file_path)
    log(f"Wrote portal results for {written} AWB(s) into the Excel file.")
    return written


# ── Selenium helpers ──────────────────────────────────────────────────────────

def login(browser, log):
    log("Navigating to login page...")
    browser.get(LOGIN_URL)
    wait = WebDriverWait(browser, 60)
    u = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#user_name")))
    browser.execute_script("arguments[0].value = arguments[1];", u, USERNAME)
    p = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#password")))
    browser.execute_script("arguments[0].value = arguments[1];", p, PASSWORD)
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#btn-login"))).click()
    wait.until(EC.url_changes(LOGIN_URL))
    log("Login successful.")


def get_span(browser, css_class):
    try:
        els = browser.find_elements(By.CSS_SELECTOR, f"span.{css_class}")
        if els:
            text = els[0].text.strip()
            return text.split(":", 1)[1].strip() if ":" in text else text
    except Exception:
        pass
    return ""


def fetch_awb(browser, prefix, serial, log):
    wait = WebDriverWait(browser, 60)
    browser.get(TRACKING_URL)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#awb_1")))
    time.sleep(1)

    inp = browser.find_element(By.CSS_SELECTOR, "#awb_1")
    browser.execute_script("arguments[0].value = arguments[1];", inp, serial)
    browser.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", inp)
    log(f"  Entered serial: {serial}")

    btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#track-submit")))
    browser.execute_script("arguments[0].click();", btn)
    log("  Waiting for results...")

    try:
        WebDriverWait(browser, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#overlay")))
    except Exception:
        pass
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#overlay")))
    time.sleep(0.5)

    origin      = get_span(browser, "tracking_origin")
    destination = get_span(browser, "tracking_destination")
    pieces      = get_span(browser, "tracking_pieces")
    weight      = get_span(browser, "tracking_weight")
    status      = get_span(browser, "tracking_status")
    flight_no   = (get_span(browser, "tracking_flight")
                   or get_span(browser, "tracking_flight_no")
                   or get_span(browser, "tracking_flightno"))
    date        = (get_span(browser, "tracking_date")
                   or get_span(browser, "tracking_flight_date"))

    return {
        "awb":         f"{prefix}-{serial}",
        "origin":      origin,
        "destination": destination,
        "pieces":      pieces,
        "weight":      weight,
        "status":      status,
        "flight_no":   flight_no,
        "date":        date,
    }


# ── Background job ────────────────────────────────────────────────────────────

def run_job(job_id, excel_path, start_row, end_row):
    job = jobs[job_id]

    def log(msg):
        job["log"].append(msg)

    output_path = os.path.join(OUTPUT_DIR, f"{job_id}_result.xlsx")
    results = []

    try:
        log(f"Scanning rows {start_row} to {end_row} for PIA AWBs...")
        target_awbs = load_awbs_from_excel(excel_path, start_row, end_row, log)

        if not target_awbs:
            log("No PIA (214-xxxx) AWBs found in the specified range.")
            job["status"] = "error"
            job["error"]  = "No PIA AWBs found in the specified row range."
            return

        log(f"Found {len(target_awbs)} PIA AWB(s). Launching headless browser...")

        options = webdriver.FirefoxOptions()
        options.add_argument("-headless")
        browser = webdriver.Firefox(options=options)

        try:
            login(browser, log)

            for i, entry in enumerate(target_awbs, 1):
                label = f"214-{entry['serial']} (row {entry['row']})"
                log(f"[{i}/{len(target_awbs)}] Fetching {label}...")
                try:
                    data = fetch_awb(browser, entry["prefix"], entry["serial"], log)
                    results.append({"status": "success", "excel_row": entry["row"], "data": data})
                    log(f"  [OK] origin={data['origin']}  dest={data['destination']}  "
                        f"pcs={data['pieces']}  wt={data['weight']}  status={data['status']}")
                except Exception as e:
                    log(f"  [FAIL] {e}")
                    results.append({
                        "awb":       f"214-{entry['serial']}",
                        "excel_row": entry["row"],
                        "status":    "error",
                        "error":     str(e),
                    })
                time.sleep(1)

        finally:
            browser.quit()
            log("Browser closed.")

        # Copy input -> output, then write results into the copy
        shutil.copy2(excel_path, output_path)
        write_results_to_excel(output_path, results, log)

        # Also save JSON alongside
        with open(os.path.join(OUTPUT_DIR, f"{job_id}_results.json"), "w") as f:
            json.dump(results, f, indent=2)

        job["result_path"] = output_path
        job["status"]      = "done"
        log("All done. Click the download button to get your updated Excel file.")

    except Exception as e:
        log(f"Fatal error: {e}")
        job["status"] = "error"
        job["error"]  = str(e)


# ── Flask routes ──────────────────────────────────────────────────────────────

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"error": str(e)}), 500


@app.route("/")
def index():
    return send_file(os.path.join(BASE_DIR, "index.html"))


@app.route("/api/start", methods=["POST"])
def api_start():
    try:
        if "excel" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["excel"]
        if not file or file.filename == "":
            return jsonify({"error": "Empty file upload"}), 400

        try:
            start_row = int(request.form.get("start_row", 2))
            end_row   = int(request.form.get("end_row", 100))
        except (TypeError, ValueError):
            return jsonify({"error": "start_row and end_row must be integers"}), 400

        if start_row < 2 or end_row < start_row:
            return jsonify({"error": "Invalid row range (start >= 2, end >= start)"}), 400

        job_id      = str(uuid.uuid4())
        upload_path = os.path.join(UPLOAD_DIR, f"{job_id}_input.xlsx")
        file.save(upload_path)

        jobs[job_id] = {"status": "running", "log": [], "result_path": None, "error": None}

        t = threading.Thread(target=run_job, args=(job_id, upload_path, start_row, end_row), daemon=True)
        t.start()

        return jsonify({"job_id": job_id})

    except Exception as e:
        return jsonify({"error": f"Unexpected error in /start: {e}"}), 500


@app.route("/api/progress/<job_id>")
def api_progress(job_id):
    def generate():
        last_idx = 0
        while True:
            job = jobs.get(job_id)
            if not job:
                yield f"data: {json.dumps({'type': 'error', 'msg': 'Job not found'})}\n\n"
                return

            new_logs = job["log"][last_idx:]
            for msg in new_logs:
                yield f"data: {json.dumps({'type': 'log', 'msg': msg})}\n\n"
            last_idx += len(new_logs)

            if job["status"] in ("done", "error"):
                yield f"data: {json.dumps({'type': 'status', 'value': job['status'], 'error': job.get('error')})}\n\n"
                return

            time.sleep(0.4)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/api/download/<job_id>")
def api_download(job_id):
    job = jobs.get(job_id)
    if not job or not job.get("result_path") or not os.path.exists(job["result_path"]):
        return "Result not available", 404
    return send_file(
        job["result_path"],
        as_attachment=True,
        download_name="awb_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    print("\nPIA AWB Tracker Web Server")
    print("Open your browser at:  http://127.0.0.1:5000\n")
    app.run(host="127.0.0.1", port=5000, threaded=True, debug=False)
